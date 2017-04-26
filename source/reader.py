import openpyxl as opx


fileName = '../excelFile/TestBook.xlsx'
sheetName = 'Sheet1'

workBook = opx.load_workbook(fileName)
sheet = workBook.get_sheet_by_name(sheetName)

maxRows = sheet.max_row
maxCol = sheet.max_column


for i in range(1, maxRows + 1):
    name = sheet.cell(row=i, column=1).value
    outputFile = open('../dump/{}.txt'.format(name), 'w')

    for j in range(1, maxCol + 1):
        outputFile.write(sheet.cell(row=i, column=j).value + '\n')

    print('Written file {}'.format(i))


print('Done Parsing')
