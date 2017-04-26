import openpyxl as excel
from generator import generate


fileName = '../excelFile/TestBook.xlsx'
sheetName = 'Sheet1'


workbook = excel.load_workbook(fileName)
sheet = workbook.get_sheet_by_name(sheetName)

sheet['A1'] = 'Name'
sheet['B1'] = 'Address'
sheet['C1'] = 'E-mail'

dataList = generate()

index = 2
for data in dataList:
    sheet.cell(row=index, column=1).value = data.getName()
    sheet.cell(row=index, column=2).value = data.getAddress()
    sheet.cell(row=index, column=3).value = data.getEmail()

    index = index + 1

workbook.save(fileName)
